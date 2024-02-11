import streamlit as st
from openpyxl import load_workbook
from openpyxl import Workbook

question = "Which of the following is integer datatype?"
st.title('TechQuest-')
em = st.text_input('E-mail')
roll = st.text_input('Roll.no')
inp = st.radio(question, ["double a", 'float a', 'char a', 'int a'])
sub = st.button('Submit')
ex = st.text_input('Enter the excel sheet:')

def mail(em):
    y = 'drngpit.ac.in'
    x = em[8:]
    if x != y:
        st.error('Enter with your college ID', icon="⚠️")
    
def rol(roll):
    if len(roll) > 7:
        st.warning('Roll number should be 7 characters')

mail(em)
rol(roll)

if sub:
    
    
    # Check if the file exists or not
    try:
        wb = load_workbook(ex)
    except FileNotFoundError:
        wb = Workbook()
        
    # Select the active worksheet or create one if it doesn't exist
    if 'Sheet' in wb.sheetnames:
        ws = wb['Sheet']
    else:
        ws = wb.active
    
    # Append data to the worksheet
    new_row = [em, roll, inp]
    ws.append(new_row)
    
    # Save the workbook
    wb.save(ex)
    
    st.success('Your answer has been submitted successfully!')
